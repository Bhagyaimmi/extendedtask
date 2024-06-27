
from django.conf import settings
from rest_framework import generics
from rest_framework.permissions import IsAuthenticated
from .models import User, Product, Article
from django.contrib import messages
from .serializers import UserSerializer, ProductSerializer, ArticleSerializer
from rest_framework import status
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from rest_framework_simplejwt.views import TokenObtainPairView
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.models import User
from django.urls import reverse_lazy
from django.utils.encoding import force_bytes
from django.utils.http import urlsafe_base64_encode
from django.template.loader import render_to_string
from django.core.mail import send_mail
from django.contrib.auth.forms import SetPasswordForm
from django.utils.http import urlsafe_base64_decode
from django.contrib.auth import get_user_model
from django.shortcuts import redirect
from django.views.generic.edit import FormView
import csv
from openpyxl import Workbook # type: ignore
from django.http import HttpResponse
from django.db.models.signals import pre_save, post_save, pre_delete, post_delete
from django.dispatch import receiver


class UserListView(generics.ListCreateAPIView):
    queryset = User.objects.all()

    serializer_class = UserSerializer
    permission_classes = [AllowAny]
    def get_queryset(self):
        queryset = User.objects.all()
        query = self.request.query_params.get('q')
        if query:
            queryset = queryset.filter(username__icontains=query)
        return queryset

class UserDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer

class ProductListView(generics.ListCreateAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]
    def get_queryset(self):
        queryset = Product.objects.all()
        query = self.request.query_params.get('q')
        if query:
            queryset = queryset.filter(name__icontains=query)
        return queryset

class ProductDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]

class ArticleListView(generics.ListCreateAPIView):
    queryset = Article.objects.all()
    serializer_class = ArticleSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Article.objects.all()
        query = self.request.query_params.get('q')
        if query:
            queryset = queryset.filter(title__icontains=query)
        return queryset

class ArticleDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Article.objects.all()
    serializer_class = ArticleSerializer
    permission_classes = [IsAuthenticated]


class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [AllowAny]

    def perform_create(self, serializer):
        user = serializer.save()

        # Send welcome email to registered users
        send_mail(
            'Welcome to Our firstprojct!',
            'Thank you for registering with us.',
            'example@example.com',
            [user.email],
            fail_silently=False,
        )
        return Response(serializer.data, status=status.HTTP_201_CREATED)

class TokenObtainPairWithUserInfoView(TokenObtainPairView):
    serializer_class = UserSerializer

class TokenRefreshView(APIView):
    def post(self, request, *args, **kwargs):
        refresh = request.data.get("refresh")
        serializer = UserSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            refresh_token = RefreshToken(refresh)
            access_token = str(refresh_token.access_token)
            response_data = {
                'access': access_token
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': 'Invalid refresh token'}, status=status.HTTP_401_UNAUTHORIZED)



class PasswordResetView(APIView):
    
    def post(self, request, *args, **kwargs):
        email = request.data.get('email')
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response({'error': 'User not found.'}, status=status.HTTP_404_NOT_FOUND)

        token = default_token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))

        reset_url = request.build_absolute_uri(reverse_lazy('password-reset-confirm', kwargs={'uidb64': uid, 'token': token}))
        message = render_to_string('email/password_reset_email.html', {'reset_url': reset_url})

        send_mail(
            'Password Reset Request',
            message,
            email,
            [email],
            fail_silently=False,
        )

        return Response({'success': 'Password reset email sent.'}, status=status.HTTP_200_OK)

class SampleEmailNotification(APIView):
    def get(self, request):
        subject = 'Sample Email Notification'
        message = 'This is a sample email notification.'
        recipient_list = ['imbhagya30@gmail.com']   #recp email
        send_mail(subject, message, settings.EMAIL_HOST_USER, recipient_list)
        return Response({'message': 'Sample email notification sent successfully.'})
    
class PasswordResetConfirmView(FormView):
    template_name = 'password_reset_confirm.html'
    form_class = SetPasswordForm
    success_url = '/password/reset/done/'


    def get_form(self, form_class=None):
        User = get_user_model()
        uidb64 = self.kwargs['uidb64']
        token = self.kwargs['token']
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)

        return self.form_class(user, self.request.POST or None)
    

    def form_valid(self, form):
        User = get_user_model()
        uidb64 = self.kwargs['uidb64']
        token = self.kwargs['token']
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)

        if default_token_generator.check_token(user, token):
            form.save()
            messages.success(self.request, 'Your password has been reset successfully.')
            return super().form_valid(form)
        else:
            messages.error(self.request, 'The password reset link is no longer valid.')
            return redirect('/password/reset/')

    def form_invalid(self, form):
        messages.error(self.request, 'There was an error resetting your password.')
        return super().form_invalid(form)


class UserCSVExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        users = User.objects.all()
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="users.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Username', 'Email'])
        for user in users:
            writer.writerow([user.id, user.username, user.email])
        return response

class UserExcelExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        users = User.objects.all()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['ID', 'Username', 'Email'])
        for user in users:
            worksheet.append([user.id, user.username, user.email])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        workbook.save(response)
        return response


class ProductCSVExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        products = Product.objects.all()
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="products.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Name', 'Description', 'Price', 'Author'])
        for product in products:
            writer.writerow([product.id, product.name, product.description, product.price, product.author.username])
        return response

class ProductExcelExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        products = Product.objects.all()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['ID', 'Name', 'Description', 'Price', 'Author'])
        for product in products:
            worksheet.append([product.id, product.name, product.description, product.price, product.author.username])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
        workbook.save(response)
        return response

class ArticleCSVExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        articles = Article.objects.all()
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="articles.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Title', 'Content', 'Author', 'Product'])
        for article in articles:
            writer.writerow([article.id, article.title, article.content, article.author.username, article.product.name if article.product else ''])
        return response

class ArticleExcelExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        articles = Article.objects.all()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['ID', 'Title', 'Content', 'Author', 'Product'])
        for article in articles:
            worksheet.append([article.id, article.title, article.content, article.author.username, article.product.name if article.product else ''])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="articles.xlsx"'
        workbook.save(response)
        return response
    
    
@receiver(pre_save, sender=Product)
def before_product_save(sender, instance, **kwargs):
    print(f'Before saving Product: {instance.name}')

@receiver(post_save, sender=Product)
def after_Product_save(sender, instance, created, **kwargs):
    if created:
        print(f'Product created: {instance.name}')
    else:
        print(f'Product updated: {instance.name}')

@receiver(pre_delete, sender=Product)
def before_Product_delete(sender, instance, **kwargs):
    print(f'Before deleting Product: {instance.name}')

@receiver(post_delete, sender=Product)
def after_Product_delete(sender, instance, **kwargs):
    print(f'Product deleted: {instance.name}')

    
@receiver(pre_save, sender=Article)
def before_Article_save(sender, instance, **kwargs):
    print(f'Before saving Article: {instance.title}')

@receiver(post_save, sender=Article)
def after_Article_save(sender, instance, created, **kwargs):
    if created:
        print(f'Article created: {instance.title}')
    else:
        print(f'Article updated: {instance.title}')

@receiver(pre_delete, sender=Article)
def before_Article_delete(sender, instance, **kwargs):
    print(f'Before deleting Article: {instance.title}')

@receiver(post_delete, sender=Article)
def after_Article_delete(sender, instance, **kwargs):
    print(f'Article deleted: {instance.title}')
